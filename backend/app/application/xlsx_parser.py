from collections.abc import Callable
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.application.column_mapping import (
    FIELD_ALIASES,
    REQUIRED_FIELDS,
    _score_header_for_field,
    normalize_header,
    resolve_sheet_field_map,
)
from app.application.csv_parser import DATE_FORMATS, _parse_amount
from app.application.errors import InvalidFileContentError
from app.application.models import NormalizedTransaction

_HEADER_SCAN_LIMIT = 20


def parse_xlsx_transactions(raw_bytes: bytes) -> list[NormalizedTransaction]:
    transactions, _field_map = parse_xlsx_transactions_with_mapping(raw_bytes=raw_bytes)
    return transactions


def parse_xlsx_transactions_with_mapping(
    raw_bytes: bytes,
    resolver: Callable[[list[str]], dict[str, str]] | None = None,
    required_fields: set[str] | None = None,
    amount_resolver: Callable[[list[object], list[str], dict[str, str]], float] | None = None,
) -> tuple[list[NormalizedTransaction], dict[str, str]]:
    field_resolver = resolver or resolve_sheet_field_map
    required = required_fields or REQUIRED_FIELDS
    resolve_amount = amount_resolver or _resolve_amount_from_amount_column
    try:
        workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl exception details are not stable.
        raise InvalidFileContentError("Unable to read XLSX content.") from exc

    if not workbook.worksheets:
        raise InvalidFileContentError("XLSX does not contain worksheets.")

    sheet = workbook.worksheets[0]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    header_row_index, fieldnames, field_map = _locate_header_row(rows, field_resolver, required)

    transactions: list[NormalizedTransaction] = []
    for raw_row in rows[header_row_index + 1 :]:
        if not raw_row:
            continue
        row_values = list(raw_row)
        if all(item is None or str(item).strip() == "" for item in row_values):
            continue

        date_raw = _require_value(row_values, fieldnames, field_map["date"], "date")
        description_raw = _require_value(row_values, fieldnames, field_map["description"], "description")
        type_raw = _get_optional_value(row_values, fieldnames, field_map.get("type"))

        amount = resolve_amount(row_values, fieldnames, field_map)
        transactions.append(
            NormalizedTransaction(
                date=_parse_date(date_raw),
                description=str(description_raw).strip(),
                amount=amount,
                type=_normalize_type(type_raw, amount),
            )
        )

    if not transactions:
        raise InvalidFileContentError("XLSX does not contain transaction rows.")
    return transactions, field_map


def _locate_header_row(
    rows: list[list[object]],
    field_resolver: Callable[[list[str]], dict[str, str]],
    required: set[str],
) -> tuple[int, list[str], dict[str, str]]:
    best_candidate: tuple[int, list[str], dict[str, str], int] | None = None
    first_error: InvalidFileContentError | None = None

    for row_index, raw_row in enumerate(rows[:_HEADER_SCAN_LIMIT]):
        fieldnames = [str(item).strip() if item is not None else "" for item in raw_row]
        if not any(fieldnames):
            continue

        try:
            field_map = field_resolver(fieldnames)
        except InvalidFileContentError as exc:
            if first_error is None:
                first_error = exc
            continue

        if required - set(field_map):
            continue

        score = _score_header_row(field_map)
        candidate = (row_index, fieldnames, field_map, score)
        if best_candidate is None or score > best_candidate[3]:
            best_candidate = candidate

    if best_candidate is None:
        if first_error is not None:
            raise first_error
        raise InvalidFileContentError("XLSX does not contain headers.")

    return best_candidate[0], best_candidate[1], best_candidate[2]


def _score_header_row(field_map: dict[str, str]) -> int:
    score = 0
    for canonical in REQUIRED_FIELDS:
        raw_header = field_map.get(canonical, "")
        score += _score_header_for_field(raw_header, FIELD_ALIASES[canonical], canonical)
    return score


def _require_value(row_values: list[object], headers: list[str], mapped_header: str, field_name: str) -> object:
    value = _extract_value(row_values, headers, mapped_header)
    if value is None or str(value).strip() == "":
        raise InvalidFileContentError(f"XLSX row has empty '{field_name}' value.")
    return value


def _get_optional_value(row_values: list[object], headers: list[str], mapped_header: str | None) -> str:
    if not mapped_header:
        return ""
    value = _extract_value(row_values, headers, mapped_header)
    return "" if value is None else str(value)


def _resolve_amount_from_amount_column(
    row_values: list[object],
    headers: list[str],
    field_map: dict[str, str],
) -> float:
    amount_raw = _require_value(row_values, headers, field_map["amount"], "amount")
    return _parse_amount(str(amount_raw))


def _extract_value(row_values: list[object], headers: list[str], mapped_header: str) -> object | None:
    try:
        index = headers.index(mapped_header)
    except ValueError as exc:
        raise InvalidFileContentError("XLSX header mapping became inconsistent.") from exc
    if index >= len(row_values):
        return None
    return row_values[index]


def _parse_date(raw: object) -> str:
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")

    value = str(raw).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    if len(value) >= 10:
        try:
            return datetime.fromisoformat(value[:10]).strftime("%Y-%m-%d")
        except ValueError:
            pass
    raise InvalidFileContentError(f"Invalid date value: {raw!r}.")


def _normalize_type(raw_type: str, amount: float) -> str:
    value = normalize_header(raw_type)
    if value in {"inflow", "credit", "entrada", "credito"}:
        return "inflow"
    if value in {"outflow", "debit", "saida", "debito"}:
        return "outflow"
    return "inflow" if amount >= 0 else "outflow"
