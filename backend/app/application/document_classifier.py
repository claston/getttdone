from __future__ import annotations

import csv
from collections.abc import Iterable
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.application.column_mapping import normalize_header

_PREVIEW_LIMIT = 20
_MAX_EVIDENCE = 6

_SEMANTIC_EXTRATO_BANCARIO = "extrato_bancario"
_SEMANTIC_CONTROLE_FINANCEIRO = "controle_financeiro"
_SEMANTIC_FLUXO_CAIXA = "fluxo_caixa"
_SEMANTIC_CONTAS_A_RECEBER = "contas_a_receber"
_SEMANTIC_CONTAS_A_PAGAR = "contas_a_pagar"
_SEMANTIC_PLANILHA_CONTABIL = "planilha_contabil_debito_credito"
_SEMANTIC_GENERICO = "generico_financeiro"

_BANK_FILENAME_HINTS = {
    "extrato",
    "statement",
    "banco",
    "bank",
    "saldo",
    "movimentacao",
    "movimentacoes",
    "lancamento",
    "lancamentos",
    "conta",
}
_ACCOUNTING_FILENAME_HINTS = {
    "contabil",
    "contabilidade",
    "debito_credito",
    "d_c",
    "diario",
    "livro_razao",
}
_FLOW_FILENAME_HINTS = {"fluxo_caixa", "fluxo", "caixa", "forecast", "projecao", "projeccao"}
_AR_FILENAME_HINTS = {"contas_a_receber", "receber", "ar", "faturas_receber", "duplicatas"}
_AP_FILENAME_HINTS = {"contas_a_pagar", "pagar", "ap", "fornecedores", "faturas_pagar"}
_CONTROL_FILENAME_HINTS = {"controle_financeiro", "financeiro", "orcamento", "budget", "gerencial"}

_BANK_HEADER_HINTS = {
    "extrato",
    "statement",
    "saldo",
    "agencia",
    "agência",
    "conta",
    "periodo",
    "período",
    "movimentacao",
    "movimentacao bancaria",
    "lancamento",
    "historico",
    "historico lancamento",
    "transacao",
    "transacao bancaria",
}
_ACCOUNTING_HEADER_HINTS = {
    "conta debito",
    "conta credito",
    "conta de debito",
    "conta de credito",
    "debit",
    "credit",
    "debito",
    "credito",
    "valor debito",
    "valor credito",
}
_AR_HEADER_HINTS = {
    "cliente",
    "receber",
    "contas a receber",
    "vencimento",
    "titulo",
    "duplicata",
    "fatura",
    "recebiveis",
}
_AP_HEADER_HINTS = {
    "fornecedor",
    "pagar",
    "contas a pagar",
    "vencimento",
    "despesa",
    "nota fiscal",
    "nf",
    "liquidacao",
}
_FLOW_HEADER_HINTS = {
    "fluxo",
    "caixa",
    "saldo inicial",
    "saldo final",
    "previsto",
    "realizado",
    "projecao",
    "forecast",
}
_CONTROL_HEADER_HINTS = {
    "controle",
    "financeiro",
    "categoria",
    "centro de custo",
    "orcamento",
    "receita",
    "despesa",
    "resumo",
    "analitico",
}

_BANK_TEXT_HINTS = {
    "pix",
    "ted",
    "doc",
    "tarifa",
    "estorno",
    "saque",
    "deposito",
    "transferencia",
    "debito automatico",
    "boleto",
    "cartao",
    "autdeb",
    "pagamento de cartao",
}
_AR_TEXT_HINTS = {
    "cliente",
    "recebimento",
    "receber",
    "fatura",
    "duplicata",
    "titulo",
    "nota fiscal",
    "vencimento",
    "cobranca",
    "boleto",
}
_AP_TEXT_HINTS = {
    "fornecedor",
    "pagamento",
    "pagar",
    "despesa",
    "nota fiscal",
    "nf",
    "vencimento",
    "fatura",
    "custo",
    "parcelamento",
}
_FLOW_TEXT_HINTS = {
    "saldo inicial",
    "saldo final",
    "previsto",
    "realizado",
    "projecao",
    "fluxo de caixa",
    "caixa",
}
_CONTROL_TEXT_HINTS = {
    "controle financeiro",
    "orcamento",
    "budget",
    "centro de custo",
    "categoria",
    "resumo",
    "analitico",
    "receita",
    "despesa",
}


@dataclass
class DocumentClassification:
    semantic_type: str
    confidence: float
    evidence: list[str] = field(default_factory=list)


def classify_document(
    filename: str,
    raw_bytes: bytes,
    *,
    layout_inference_name: str | None = None,
    layout_inference_confidence: float | None = None,
) -> DocumentClassification:
    extension = Path(filename).suffix.lower().lstrip(".")
    profile = _build_profile(filename=filename, raw_bytes=raw_bytes, extension=extension)
    scores: dict[str, float] = {
        _SEMANTIC_EXTRATO_BANCARIO: 0.0,
        _SEMANTIC_CONTROLE_FINANCEIRO: 0.0,
        _SEMANTIC_FLUXO_CAIXA: 0.0,
        _SEMANTIC_CONTAS_A_RECEBER: 0.0,
        _SEMANTIC_CONTAS_A_PAGAR: 0.0,
        _SEMANTIC_PLANILHA_CONTABIL: 0.0,
        _SEMANTIC_GENERICO: 0.0,
    }
    evidence: dict[str, list[str]] = {key: [] for key in scores}

    _score_filename(profile.filename_normalized, scores, evidence)
    _score_headers(profile.header_text, scores, evidence)
    _score_body(profile.body_text, scores, evidence)
    _score_structure(profile, scores, evidence)

    if extension in {"pdf", "ofx"}:
        _add_score(scores, evidence, _SEMANTIC_EXTRATO_BANCARIO, 40.0, f"file extension suggests bank statement: {extension}")

    if layout_inference_name:
        layout_name = normalize_header(layout_inference_name)
        if "statement" in layout_name:
            _add_score(
                scores,
                evidence,
                _SEMANTIC_EXTRATO_BANCARIO,
                35.0,
                f"pdf layout inference suggests bank statement: {layout_inference_name}",
            )
        if layout_inference_confidence is not None and layout_inference_confidence >= 0.85:
            _add_score(
                scores,
                evidence,
                _SEMANTIC_EXTRATO_BANCARIO,
                10.0,
                f"high pdf layout confidence: {layout_inference_confidence:.2f}",
            )

    _maybe_add_generic_fallback(profile, scores, evidence)

    ordered = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    best_type, best_score = ordered[0]
    second_score = ordered[1][1] if len(ordered) > 1 else 0.0

    if best_score <= 0:
        return DocumentClassification(
            semantic_type=_SEMANTIC_GENERICO,
            confidence=0.0,
            evidence=[],
        )

    confidence = _confidence_from_scores(best_score, second_score)
    top_evidence = evidence.get(best_type, [])[:_MAX_EVIDENCE]
    if best_type == _SEMANTIC_GENERICO and confidence < 0.35:
        return DocumentClassification(
            semantic_type=_SEMANTIC_GENERICO,
            confidence=round(max(confidence, 0.25), 2),
            evidence=top_evidence or ["no strong semantic signals found"],
        )

    return DocumentClassification(
        semantic_type=best_type,
        confidence=confidence,
        evidence=top_evidence,
    )


@dataclass
class _DocumentProfile:
    filename_normalized: str
    header_text: str
    body_text: str
    row_count: int
    has_headers: bool
    extension: str
    sample_rows: list[list[str]]


def _build_profile(filename: str, raw_bytes: bytes, extension: str) -> _DocumentProfile:
    filename_normalized = normalize_header(Path(filename).stem)
    if extension == "csv":
        return _build_csv_profile(filename_normalized, raw_bytes, extension)
    if extension == "xlsx":
        return _build_xlsx_profile(filename_normalized, raw_bytes, extension)
    text = _decode_text_preview(raw_bytes)
    return _DocumentProfile(
        filename_normalized=filename_normalized,
        header_text="",
        body_text=text,
        row_count=0,
        has_headers=False,
        extension=extension,
        sample_rows=[],
    )


def _build_csv_profile(filename_normalized: str, raw_bytes: bytes, extension: str) -> _DocumentProfile:
    text = _decode_text_preview(raw_bytes)
    delimiter = _detect_delimiter(text)
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = _take_rows(reader, _PREVIEW_LIMIT)
    header = rows[0] if rows else []
    body_rows = rows[1:] if len(rows) > 1 else []
    return _DocumentProfile(
        filename_normalized=filename_normalized,
        header_text=_join_tokens(header),
        body_text=_join_tokens(row for row in body_rows),
        row_count=max(len(rows) - 1, 0),
        has_headers=bool(header),
        extension=extension,
        sample_rows=rows,
    )


def _build_xlsx_profile(filename_normalized: str, raw_bytes: bytes, extension: str) -> _DocumentProfile:
    try:
        workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception:
        return _DocumentProfile(
            filename_normalized=filename_normalized,
            header_text="",
            body_text="",
            row_count=0,
            has_headers=False,
            extension=extension,
            sample_rows=[],
        )

    if not workbook.worksheets:
        return _DocumentProfile(
            filename_normalized=filename_normalized,
            header_text="",
            body_text="",
            row_count=0,
            has_headers=False,
            extension=extension,
            sample_rows=[],
        )

    rows = [
        ["" if value is None else str(value) for value in row]
        for row in workbook.worksheets[0].iter_rows(values_only=True)
    ][: _PREVIEW_LIMIT]
    header = rows[0] if rows else []
    body_rows = rows[1:] if len(rows) > 1 else []
    return _DocumentProfile(
        filename_normalized=filename_normalized,
        header_text=_join_tokens(header),
        body_text=_join_tokens(row for row in body_rows),
        row_count=max(len(rows) - 1, 0),
        has_headers=bool(header),
        extension=extension,
        sample_rows=rows,
    )


def _score_filename(filename_normalized: str, scores: dict[str, float], evidence: dict[str, list[str]]) -> None:
    _score_terms(filename_normalized, scores, evidence, _SEMANTIC_EXTRATO_BANCARIO, _BANK_FILENAME_HINTS, 18.0, "filename")
    _score_terms(filename_normalized, scores, evidence, _SEMANTIC_PLANILHA_CONTABIL, _ACCOUNTING_FILENAME_HINTS, 18.0, "filename")
    _score_terms(filename_normalized, scores, evidence, _SEMANTIC_FLUXO_CAIXA, _FLOW_FILENAME_HINTS, 15.0, "filename")
    _score_terms(filename_normalized, scores, evidence, _SEMANTIC_CONTAS_A_RECEBER, _AR_FILENAME_HINTS, 15.0, "filename")
    _score_terms(filename_normalized, scores, evidence, _SEMANTIC_CONTAS_A_PAGAR, _AP_FILENAME_HINTS, 15.0, "filename")
    _score_terms(filename_normalized, scores, evidence, _SEMANTIC_CONTROLE_FINANCEIRO, _CONTROL_FILENAME_HINTS, 12.0, "filename")


def _score_headers(header_text: str, scores: dict[str, float], evidence: dict[str, list[str]]) -> None:
    _score_terms(header_text, scores, evidence, _SEMANTIC_EXTRATO_BANCARIO, _BANK_HEADER_HINTS, 12.0, "headers")
    _score_terms(header_text, scores, evidence, _SEMANTIC_PLANILHA_CONTABIL, _ACCOUNTING_HEADER_HINTS, 25.0, "headers")
    _score_terms(header_text, scores, evidence, _SEMANTIC_CONTAS_A_RECEBER, _AR_HEADER_HINTS, 10.0, "headers")
    _score_terms(header_text, scores, evidence, _SEMANTIC_CONTAS_A_PAGAR, _AP_HEADER_HINTS, 10.0, "headers")
    _score_terms(header_text, scores, evidence, _SEMANTIC_FLUXO_CAIXA, _FLOW_HEADER_HINTS, 10.0, "headers")
    _score_terms(header_text, scores, evidence, _SEMANTIC_CONTROLE_FINANCEIRO, _CONTROL_HEADER_HINTS, 8.0, "headers")


def _score_body(body_text: str, scores: dict[str, float], evidence: dict[str, list[str]]) -> None:
    _score_terms(body_text, scores, evidence, _SEMANTIC_EXTRATO_BANCARIO, _BANK_TEXT_HINTS, 8.0, "rows")
    _score_terms(body_text, scores, evidence, _SEMANTIC_CONTAS_A_RECEBER, _AR_TEXT_HINTS, 8.0, "rows")
    _score_terms(body_text, scores, evidence, _SEMANTIC_CONTAS_A_PAGAR, _AP_TEXT_HINTS, 8.0, "rows")
    _score_terms(body_text, scores, evidence, _SEMANTIC_FLUXO_CAIXA, _FLOW_TEXT_HINTS, 8.0, "rows")
    _score_terms(body_text, scores, evidence, _SEMANTIC_CONTROLE_FINANCEIRO, _CONTROL_TEXT_HINTS, 6.0, "rows")


def _score_structure(profile: _DocumentProfile, scores: dict[str, float], evidence: dict[str, list[str]]) -> None:
    header_text = profile.header_text
    normalized_rows = [" ".join(row) for row in profile.sample_rows[1:]]
    joined_rows = _normalize_joined(normalized_rows)

    if _contains_any(header_text, {"debit credit", "conta debito", "conta credito", "vlr debito", "vlr credito"}):
        _add_score(
            scores,
            evidence,
            _SEMANTIC_PLANILHA_CONTABIL,
            120.0,
            "split debit/credit columns detected",
        )

    if _contains_any(header_text, {"saldo inicial", "saldo final"}):
        _add_score(scores, evidence, _SEMANTIC_FLUXO_CAIXA, 20.0, "opening/closing balance columns detected")

    if _contains_any(joined_rows, {"saldo inicial", "saldo final", "previsto", "realizado"}):
        _add_score(scores, evidence, _SEMANTIC_FLUXO_CAIXA, 16.0, "cash flow planning terms found in rows")

    if _contains_any(joined_rows, {"pix", "ted", "doc", "tarifa", "estorno", "transferencia"}):
        _add_score(scores, evidence, _SEMANTIC_EXTRATO_BANCARIO, 14.0, "bank transaction terms found in rows")

    if _contains_any(joined_rows, {"cliente", "recebimento", "fatura", "duplicata", "titulo"}):
        _add_score(scores, evidence, _SEMANTIC_CONTAS_A_RECEBER, 14.0, "accounts receivable terms found in rows")

    if _contains_any(joined_rows, {"fornecedor", "pagamento", "nota fiscal", "nf", "vencimento"}):
        _add_score(scores, evidence, _SEMANTIC_CONTAS_A_PAGAR, 14.0, "accounts payable terms found in rows")

    if _contains_any(joined_rows, {"centro de custo", "categoria", "orcamento", "budget", "resumo", "analitico"}):
        _add_score(scores, evidence, _SEMANTIC_CONTROLE_FINANCEIRO, 12.0, "financial control terms found in rows")

    if profile.extension in {"pdf", "ofx"}:
        _add_score(scores, evidence, _SEMANTIC_EXTRATO_BANCARIO, 18.0, f"{profile.extension.upper()} is usually a bank statement source")


def _maybe_add_generic_fallback(profile: _DocumentProfile, scores: dict[str, float], evidence: dict[str, list[str]]) -> None:
    if profile.row_count <= 0:
        return

    generic_signals = 0.0
    if profile.has_headers and _contains_any(profile.header_text, {"date", "data", "description", "descricao", "amount", "valor"}):
        generic_signals += 8.0
    if _contains_any(profile.body_text, {"receita", "despesa", "saldo", "lancamento", "movimento"}):
        generic_signals += 4.0

    if generic_signals > 0:
        _add_score(scores, evidence, _SEMANTIC_GENERICO, generic_signals, "generic financial rows and columns detected")


def _score_terms(
    text: str,
    scores: dict[str, float],
    evidence: dict[str, list[str]],
    semantic_type: str,
    terms: Iterable[str],
    weight: float,
    source: str,
) -> None:
    hits = [term for term in terms if _contains_any(text, {term})]
    if not hits:
        return

    _add_score(
        scores,
        evidence,
        semantic_type,
        weight + max(0.0, (len(hits) - 1) * (weight / 3.0)),
        f"{source} signals: {', '.join(hits[:3])}",
    )


def _add_score(
    scores: dict[str, float],
    evidence: dict[str, list[str]],
    semantic_type: str,
    score: float,
    message: str,
) -> None:
    scores[semantic_type] = scores.get(semantic_type, 0.0) + score
    bucket = evidence.setdefault(semantic_type, [])
    if len(bucket) < _MAX_EVIDENCE:
        bucket.append(message)


def _confidence_from_scores(best_score: float, second_score: float) -> float:
    if best_score <= 0:
        return 0.0
    confidence = best_score / (best_score + second_score + 10.0)
    return round(min(0.99, max(0.0, confidence)), 2)


def _contains_any(text: str, terms: Iterable[str]) -> bool:
    normalized = _normalize_joined([text])
    return any(_contains_text(normalized, term) for term in terms)


def _contains_text(text: str, term: str) -> bool:
    normalized_term = normalize_header(term)
    if not normalized_term:
        return False
    return normalized_term in text


def _normalize_joined(parts: Iterable[str]) -> str:
    tokens = [normalize_header(part) for part in parts if part is not None]
    return " ".join(token for token in tokens if token)


def _join_tokens(rows: Iterable[Iterable[str]]) -> str:
    flattened: list[str] = []
    for row in rows:
        if isinstance(row, str):
            items: Iterable[str] = [row]
        else:
            items = row
        flattened.extend(str(item) for item in items if item is not None and str(item).strip())
    return _normalize_joined(flattened)


def _decode_text_preview(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:5])
    if not sample:
        return ","
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        return ";" if sample.count(";") > sample.count(",") else ","


def _take_rows(reader: csv.reader[Any], limit: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for idx, row in enumerate(reader):
        if idx >= limit:
            break
        rows.append(["" if item is None else str(item) for item in row])
    return rows
