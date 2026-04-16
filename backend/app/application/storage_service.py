import csv
import json
from dataclasses import asdict
from datetime import datetime, timedelta, timezone
from io import StringIO
from pathlib import Path
from typing import Callable
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.application.errors import AnalysisAccessDeniedError, AnalysisEditConflictError, AnalysisNotFoundError
from app.application.models import AnalysisData, NormalizedTransaction, TransactionRow
from app.application.ofx_writer import build_ofx_statement


class TempAnalysisStorage:
    def __init__(
        self,
        root_dir: Path,
        ttl_seconds: int = 24 * 60 * 60,
        now_provider: Callable[[], datetime] | None = None,
    ) -> None:
        self.root_dir = root_dir
        self.ttl_seconds = ttl_seconds
        self.now_provider = now_provider or (lambda: datetime.now(timezone.utc))
        self.root_dir.mkdir(parents=True, exist_ok=True)

    def save_analysis(self, data: AnalysisData) -> str:
        analysis_dir = self.root_dir / data.analysis_id
        analysis_dir.mkdir(parents=True, exist_ok=True)
        now = self.now_provider()
        expires_at = now + timedelta(seconds=self.ttl_seconds)
        report_rows = data.report_transactions or data.preview_transactions
        updated_at = data.updated_at or now.isoformat()

        json_path = analysis_dir / "analysis.json"
        json_path.write_text(
            json.dumps(
                {
                    **asdict(data),
                    "created_at": now.isoformat(),
                    "updated_at": updated_at,
                    "expires_at": expires_at.isoformat(),
                    "preview_transactions": [asdict(item) for item in data.preview_transactions],
                    "report_transactions": [asdict(item) for item in report_rows],
                },
                ensure_ascii=True,
                indent=2,
            ),
            encoding="utf-8",
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transacoes"
        sheet.append(["date", "description", "amount", "category", "reconciliation_status"])
        for item in report_rows:
            sheet.append([item.date, item.description, item.amount, item.category, item.reconciliation_status])
        self._format_transacoes_sheet(sheet)
        self._add_conciliacao_sheet(workbook, data)
        workbook.save(analysis_dir / "report.xlsx")
        self._write_convert_artifacts(analysis_dir, report_rows)
        return expires_at.isoformat()

    def get_report_path(self, analysis_id: str) -> Path:
        analysis_dir = self.root_dir / analysis_id
        report_path = analysis_dir / "report.xlsx"
        if not report_path.exists():
            raise AnalysisNotFoundError
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError
        return report_path

    def get_convert_report_path(self, analysis_id: str, file_format: str) -> Path:
        analysis_dir = self.root_dir / analysis_id
        suffix = "ofx" if file_format == "ofx" else "csv"
        report_path = analysis_dir / f"converted.{suffix}"
        if not report_path.exists():
            raise AnalysisNotFoundError
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError
        return report_path

    def get_upload_filename(self, analysis_id: str) -> str | None:
        analysis_dir = self.root_dir / analysis_id
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError

        metadata_path = analysis_dir / "analysis.json"
        if not metadata_path.exists():
            raise AnalysisNotFoundError
        try:
            content = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return None

        raw_name = content.get("upload_filename")
        if not isinstance(raw_name, str):
            return None
        name = raw_name.strip()
        return name or None

    def set_convert_owner(self, analysis_id: str, identity_type: str, identity_id: str) -> None:
        analysis_dir = self.root_dir / analysis_id
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError

        metadata_path = analysis_dir / "analysis.json"
        if not metadata_path.exists():
            raise AnalysisNotFoundError
        try:
            content = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            raise AnalysisNotFoundError from None

        owner_type = str(content.get("owner_identity_type") or "").strip()
        owner_id = str(content.get("owner_identity_id") or "").strip()
        if owner_type and owner_id:
            if owner_type != identity_type or owner_id != identity_id:
                raise AnalysisAccessDeniedError
            return

        content["owner_identity_type"] = identity_type
        content["owner_identity_id"] = identity_id
        metadata_path.write_text(json.dumps(content, ensure_ascii=True, indent=2), encoding="utf-8")

    def assert_convert_owner(self, analysis_id: str, identity_type: str, identity_id: str) -> None:
        analysis_dir = self.root_dir / analysis_id
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError

        metadata_path = analysis_dir / "analysis.json"
        if not metadata_path.exists():
            raise AnalysisNotFoundError
        try:
            content = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            raise AnalysisNotFoundError from None

        owner_type = str(content.get("owner_identity_type") or "").strip()
        owner_id = str(content.get("owner_identity_id") or "").strip()
        if not owner_type and not owner_id:
            return
        if owner_type != identity_type or owner_id != identity_id:
            raise AnalysisAccessDeniedError

    def apply_convert_edits(
        self,
        analysis_id: str,
        edits: list[dict[str, object]],
        expected_updated_at: str | None = None,
    ) -> dict[str, object]:
        analysis_dir = self.root_dir / analysis_id
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError

        metadata_path = analysis_dir / "analysis.json"
        if not metadata_path.exists():
            raise AnalysisNotFoundError

        try:
            content = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            raise AnalysisNotFoundError from None

        preview_rows = self._parse_transaction_rows(content.get("preview_transactions", []))
        report_rows_raw = content.get("report_transactions") or content.get("preview_transactions", [])
        report_rows = self._parse_transaction_rows(report_rows_raw)
        current_updated_at = str(content.get("updated_at") or "").strip()

        if len(preview_rows) == 0 or len(report_rows) == 0:
            raise AnalysisNotFoundError
        if expected_updated_at and current_updated_at and expected_updated_at != current_updated_at:
            raise AnalysisEditConflictError

        for edit in edits:
            row_id_raw = edit.get("row_id")
            row_index = self._resolve_row_index(row_id_raw, max_index=len(preview_rows))
            amount = self._build_amount_from_credit_debit(edit.get("credit"), edit.get("debit"))
            description = str(edit.get("description") or "").strip()
            date = str(edit.get("date") or "").strip()
            if not description or not date:
                raise ValueError("Date and description are required.")

            report_target = report_rows[row_index]
            report_target.date = date
            report_target.description = description
            report_target.amount = amount

            preview_target = preview_rows[row_index]
            preview_target.date = date
            preview_target.description = description
            preview_target.amount = amount

        total_inflows = round(sum(item.amount for item in report_rows if item.amount > 0), 2)
        total_outflows = round(sum(item.amount for item in report_rows if item.amount < 0), 2)
        net_total = round(total_inflows + total_outflows, 2)

        content["transactions_total"] = len(report_rows)
        content["total_inflows"] = total_inflows
        content["total_outflows"] = total_outflows
        content["net_total"] = net_total
        content["preview_transactions"] = [asdict(item) for item in preview_rows]
        content["report_transactions"] = [asdict(item) for item in report_rows]
        new_updated_at = self.now_provider().isoformat()
        content["updated_at"] = new_updated_at

        metadata_path.write_text(json.dumps(content, ensure_ascii=True, indent=2), encoding="utf-8")
        self._write_report_workbook(analysis_dir, content=content, report_rows=report_rows, preview_rows=preview_rows)
        self._write_convert_artifacts(analysis_dir, report_rows)

        return {
            "processing_id": analysis_id,
            "transactions_total": len(report_rows),
            "total_inflows": total_inflows,
            "total_outflows": total_outflows,
            "net_total": net_total,
            "preview_transactions": [asdict(item) for item in preview_rows],
            "updated_at": new_updated_at,
        }

    def save_reconcile_report(
        self,
        summary: dict[str, int],
        reconciliation_rows: list[dict[str, str | float | None]],
        problems: list[dict[str, str]],
    ) -> tuple[str, str]:
        analysis_id = self._build_analysis_id(prefix="rc")
        analysis_dir = self.root_dir / analysis_id
        analysis_dir.mkdir(parents=True, exist_ok=True)
        now = self.now_provider()
        expires_at = now + timedelta(seconds=self.ttl_seconds)

        json_path = analysis_dir / "reconcile.json"
        json_path.write_text(
            json.dumps(
                {
                    "analysis_id": analysis_id,
                    "created_at": now.isoformat(),
                    "expires_at": expires_at.isoformat(),
                    "summary": summary,
                    "reconciliation_rows": reconciliation_rows,
                    "problems": problems,
                },
                ensure_ascii=True,
                indent=2,
            ),
            encoding="utf-8",
        )

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumo"
        summary_sheet.append(["Metrica", "Valor"])
        summary_sheet.append(["Total linhas extrato", summary.get("total_bank_rows", 0)])
        summary_sheet.append(["Total linhas planilha", summary.get("total_sheet_rows", 0)])
        summary_sheet.append(["Conciliados", summary.get("conciliated_count", 0)])
        summary_sheet.append(["Pendentes", summary.get("pending_count", 0)])
        summary_sheet.append(["Divergentes", summary.get("divergent_count", 0)])
        self._format_transacoes_sheet(summary_sheet)

        detail_sheet = workbook.create_sheet(title="Conciliacao_Detalhada")
        detail_headers = [
            "Linha",
            "Fonte",
            "Data",
            "Descricao",
            "Valor",
            "Status",
            "Regra de match",
            "Motivo",
            "Linha pareada",
        ]
        detail_sheet.append(detail_headers)
        for row in reconciliation_rows:
            detail_sheet.append(
                [
                    row.get("row_id"),
                    self._translate_source_label(row.get("source")),
                    row.get("date"),
                    row.get("description"),
                    row.get("amount"),
                    self._translate_status_label(row.get("status")),
                    self._translate_match_rule_label(row.get("match_rule")),
                    self._translate_reason_label(row.get("reason")),
                    row.get("matched_row_id"),
                ]
            )
        self._format_transacoes_sheet(detail_sheet)

        problems_sheet = workbook.create_sheet(title="Problemas")
        problem_row_headers = [
            "Linha",
            "Fonte",
            "Data",
            "Descricao",
            "Valor",
            "Status",
            "Motivo",
            "Linha pareada",
        ]
        problems_sheet.append(problem_row_headers)
        problematic_rows = [
            row for row in reconciliation_rows if row.get("status") in {"pendente", "divergente"}
        ]
        for row in problematic_rows:
            problems_sheet.append(
                [
                    row.get("row_id"),
                    self._translate_source_label(row.get("source")),
                    row.get("date"),
                    row.get("description"),
                    row.get("amount"),
                    self._translate_status_label(row.get("status")),
                    self._translate_reason_label(row.get("reason")),
                    row.get("matched_row_id"),
                ]
            )
        if len(problematic_rows) == 0:
            problems_sheet.append(
                [
                    "nenhum",
                    "Sistema",
                    "",
                    "Nenhuma pendencia ou divergencia foi identificada.",
                    "",
                    "-",
                    "-",
                    "",
                ]
            )
        self._format_transacoes_sheet(problems_sheet)

        workbook.save(analysis_dir / "reconcile_report.xlsx")

        csv_headers = detail_headers
        csv_lines = [",".join(csv_headers)]
        for row in reconciliation_rows:
            values = [
                self._escape_csv_value(value)
                for value in [
                    row.get("row_id"),
                    self._translate_source_label(row.get("source")),
                    row.get("date"),
                    row.get("description"),
                    row.get("amount"),
                    self._translate_status_label(row.get("status")),
                    self._translate_match_rule_label(row.get("match_rule")),
                    self._translate_reason_label(row.get("reason")),
                    row.get("matched_row_id"),
                ]
            ]
            csv_lines.append(",".join(values))
        (analysis_dir / "reconcile_report.csv").write_text("\n".join(csv_lines), encoding="utf-8")

        return analysis_id, expires_at.isoformat()

    def get_reconcile_report_path(self, analysis_id: str, file_format: str) -> Path:
        analysis_dir = self.root_dir / analysis_id
        if self._is_expired(analysis_dir):
            self._cleanup_analysis(analysis_dir)
            raise AnalysisNotFoundError

        suffix = "xlsx" if file_format == "xlsx" else "csv"
        report_path = analysis_dir / f"reconcile_report.{suffix}"
        if not report_path.exists():
            raise AnalysisNotFoundError
        return report_path

    def _write_convert_artifacts(self, analysis_dir: Path, report_rows: list[TransactionRow]) -> None:
        normalized_transactions = [
            NormalizedTransaction(
                date=item.date,
                description=item.description,
                amount=item.amount,
                type="inflow" if item.amount >= 0 else "outflow",
            )
            for item in report_rows
        ]

        (analysis_dir / "converted.ofx").write_text(
            build_ofx_statement(normalized_transactions),
            encoding="utf-8",
        )

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(["date", "description", "amount", "category", "reconciliation_status"])
        for item in report_rows:
            writer.writerow([item.date, item.description, item.amount, item.category, item.reconciliation_status])
        (analysis_dir / "converted.csv").write_text(csv_buffer.getvalue(), encoding="utf-8")

    def _write_report_workbook(
        self,
        analysis_dir: Path,
        content: dict[str, object],
        report_rows: list[TransactionRow],
        preview_rows: list[TransactionRow],
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transacoes"
        sheet.append(["date", "description", "amount", "category", "reconciliation_status"])
        for item in report_rows:
            sheet.append([item.date, item.description, item.amount, item.category, item.reconciliation_status])
        self._format_transacoes_sheet(sheet)

        snapshot = AnalysisData(
            analysis_id=str(content.get("analysis_id", "")),
            file_type=str(content.get("file_type", "")),
            upload_filename=content.get("upload_filename") if isinstance(content.get("upload_filename"), str) else None,
            transactions_total=int(content.get("transactions_total", len(report_rows))),
            total_inflows=float(content.get("total_inflows", 0.0)),
            total_outflows=float(content.get("total_outflows", 0.0)),
            net_total=float(content.get("net_total", 0.0)),
            preview_transactions=preview_rows,
            report_transactions=report_rows,
            matched_groups=int(content.get("matched_groups", 0)),
            reversed_entries=int(content.get("reversed_entries", 0)),
            potential_duplicates=int(content.get("potential_duplicates", 0)),
        )
        self._add_conciliacao_sheet(workbook, snapshot)
        workbook.save(analysis_dir / "report.xlsx")

    def _parse_transaction_rows(self, rows_raw: object) -> list[TransactionRow]:
        parsed: list[TransactionRow] = []
        if not isinstance(rows_raw, list):
            return parsed
        for item in rows_raw:
            if not isinstance(item, dict):
                continue
            parsed.append(
                TransactionRow(
                    date=str(item.get("date") or ""),
                    description=str(item.get("description") or ""),
                    amount=float(item.get("amount") or 0.0),
                    category=str(item.get("category") or "Outros"),
                    reconciliation_status=str(item.get("reconciliation_status") or "unmatched"),
                )
            )
        return parsed

    def _resolve_row_index(self, row_id_raw: object, max_index: int) -> int:
        row_id = str(row_id_raw or "").strip()
        if not row_id.startswith("row_"):
            raise ValueError("Invalid row_id.")
        suffix = row_id.split("_", 1)[1]
        if not suffix.isdigit():
            raise ValueError("Invalid row_id.")
        idx = int(suffix) - 1
        if idx < 0 or idx >= max_index:
            raise ValueError("row_id out of bounds.")
        return idx

    def _build_amount_from_credit_debit(self, credit_raw: object, debit_raw: object) -> float:
        credit = None if credit_raw is None else float(credit_raw)
        debit = None if debit_raw is None else float(debit_raw)
        has_credit = credit is not None and credit > 0
        has_debit = debit is not None and debit > 0
        if has_credit == has_debit:
            raise ValueError("Provide only one of credit or debit.")
        if has_credit:
            return round(float(credit), 2)
        return -round(abs(float(debit)), 2)

    def _is_expired(self, analysis_dir: Path) -> bool:
        metadata_path = analysis_dir / "analysis.json"
        try:
            content = json.loads(metadata_path.read_text(encoding="utf-8"))
            expires_at_raw = content.get("expires_at")
            if not isinstance(expires_at_raw, str):
                return False
            expires_at = datetime.fromisoformat(expires_at_raw)
            return self.now_provider() > expires_at
        except (OSError, TypeError, json.JSONDecodeError, ValueError):
            return False

    def _cleanup_analysis(self, analysis_dir: Path) -> None:
        if not analysis_dir.exists():
            return
        for file in analysis_dir.glob("**/*"):
            if file.is_file():
                file.unlink(missing_ok=True)
        for directory in sorted((item for item in analysis_dir.glob("**/*") if item.is_dir()), key=lambda x: len(x.parts), reverse=True):
            directory.rmdir()
        analysis_dir.rmdir()

    def _format_transacoes_sheet(self, sheet) -> None:
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        for column_index in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            max_len = 0
            for row_index in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_index, column=column_index).value
                max_len = max(max_len, len(str(cell_value)) if cell_value is not None else 0)
            sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 80)

    def _add_conciliacao_sheet(self, workbook: Workbook, data: AnalysisData) -> None:
        sheet = workbook.create_sheet(title="Conciliacao")
        sheet.append(["metric", "value"])
        sheet.append(["matched_groups", data.matched_groups])
        sheet.append(["reversed_entries", data.reversed_entries])
        sheet.append(["potential_duplicates", data.potential_duplicates])
        sheet.append([])
        sheet.append(["date", "description", "amount", "category", "reconciliation_status"])
        for item in data.preview_transactions:
            if item.reconciliation_status != "unmatched":
                sheet.append([item.date, item.description, item.amount, item.category, item.reconciliation_status])

        if sheet.max_row == 6:
            sheet.append(["-", "No reconciled entries in preview", "", "", "unmatched"])

        self._format_transacoes_sheet(sheet)

    def _build_analysis_id(self, prefix: str) -> str:
        return f"{prefix}_{uuid4().hex[:12]}"

    def _escape_csv_value(self, value: object) -> str:
        text = "" if value is None else str(value)
        if any(char in text for char in [",", "\"", "\n", "\r"]):
            return f"\"{text.replace('\"', '\"\"')}\""
        return text

    def _translate_source_label(self, source: object) -> str:
        text = "" if source is None else str(source)
        labels = {
            "bank": "Banco",
            "sheet": "Planilha",
            "system": "Sistema",
        }
        return labels.get(text, text or "-")

    def _translate_status_label(self, status: object) -> str:
        text = "" if status is None else str(status)
        labels = {
            "conciliado": "Conciliado",
            "pendente": "Pendente",
            "divergente": "Divergente",
            "none": "-",
        }
        return labels.get(text, text or "-")

    def _translate_match_rule_label(self, match_rule: object) -> str:
        text = "" if match_rule is None else str(match_rule)
        labels = {
            "exact": "Exato",
            "date_tolerance": "Tolerancia de data",
            "description_similarity": "Similaridade de descricao",
            "none": "-",
        }
        return labels.get(text, text or "-")

    def _translate_reason_label(self, reason: object) -> str:
        text = "" if reason is None else str(reason)
        labels = {
            "missing_in_sheet": "Pendente na planilha",
            "missing_in_bank": "Pendente no banco",
            "amount_mismatch": "Diferenca de valor",
            "date_out_of_tolerance_window": "Data fora da tolerancia",
            "matched_equal_amount_same_day": "Valor igual na mesma data",
            "matched_equal_amount_within_2_days": "Valor igual dentro de 2 dias",
            "matched_equal_amount_with_similar_description": "Valor igual com descricao similar",
            "none": "-",
        }
        return labels.get(text, text or "-")
