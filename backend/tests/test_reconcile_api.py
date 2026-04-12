from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app


def _build_xlsx_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_reconcile_happy_path_accepts_bank_and_sheet_files() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,teste,-100", "text/csv"),
            "sheet_file": ("sheet.csv", b"data,valor,descricao\n2026-04-01,100,TEST", "text/csv"),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["analysis_id"].startswith("rc_")
    assert payload["status"] == "accepted"
    assert payload["bank_filename"] == "bank.csv"
    assert payload["bank_file_type"] == "csv"
    assert payload["sheet_filename"] == "sheet.csv"
    assert payload["sheet_file_type"] == "csv"
    assert payload["bank_rows_parsed"] == 1
    assert payload["sheet_rows_parsed"] == 1
    assert payload["sheet_mapping_detected"] == {
        "date": "data",
        "amount": "valor",
        "description": "descricao",
    }
    assert payload["exact_matches_count"] == 0
    assert payload["date_tolerance_matches_count"] == 0
    assert payload["description_similarity_matches_count"] == 0
    assert payload["total_matches_count"] == 0
    assert payload["conciliated_count"] == 0
    assert payload["pending_count"] == 0
    assert payload["divergent_count"] == 2
    assert payload["bank_unmatched_count"] == 1
    assert payload["sheet_unmatched_count"] == 1
    assert payload["exact_matches_preview"] == []
    assert payload["date_tolerance_matches_preview"] == []
    assert payload["description_similarity_matches_preview"] == []
    assert len(payload["reconciliation_rows"]) == 2
    assert payload["reconciliation_rows"][0]["status"] == "divergente"
    assert payload["summary"] == {
        "total_bank_rows": 1,
        "total_sheet_rows": 1,
        "conciliated_count": 0,
        "pending_count": 0,
        "divergent_count": 2,
    }
    assert isinstance(payload["problems"], list)
    assert len(payload["normalization_preview"]) == 2
    assert payload["normalization_preview"][0]["source"] == "bank"
    assert payload["normalization_preview"][1]["source"] == "sheet"
    assert payload["expires_at"] is not None


def test_reconcile_rejects_unsupported_bank_file_type() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.pdf", b"%PDF", "application/pdf"),
            "sheet_file": ("sheet.xlsx", b"fake-xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Unsupported bank file type. Use CSV, XLSX, or OFX."


def test_reconcile_rejects_unsupported_sheet_file_type() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,TEST,10", "text/csv"),
            "sheet_file": ("sheet.ofx", b"<OFX>...</OFX>", "application/octet-stream"),
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Unsupported sheet file type. Use CSV or XLSX."


def test_reconcile_returns_422_when_sheet_is_missing_required_columns() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,TEST,10", "text/csv"),
            "sheet_file": ("sheet.csv", b"descricao,valor\nTEST,100", "text/csv"),
        },
    )

    assert response.status_code == 422
    assert "missing required columns" in response.json()["detail"]


def test_reconcile_accepts_sheet_xlsx_with_alias_columns() -> None:
    client = TestClient(app)
    sheet_raw = _build_xlsx_bytes(
        [
            ["dt_lancamento", "vlr", "historico"],
            ["01/04/2026", "1200,00", "RECEBIMENTO CLIENTE"],
            ["02/04/2026", "-250,50", "PAGAMENTO FORNECEDOR"],
        ]
    )

    response = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,TEST,10", "text/csv"),
            "sheet_file": (
                "sheet.xlsx",
                sheet_raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["sheet_file_type"] == "xlsx"
    assert payload["bank_rows_parsed"] == 1
    assert payload["sheet_rows_parsed"] == 2
    assert payload["sheet_mapping_detected"] == {
        "date": "dt_lancamento",
        "amount": "vlr",
        "description": "historico",
    }
    assert len(payload["normalization_preview"]) == 3


def test_reconcile_returns_422_for_ambiguous_sheet_mapping() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,TEST,10", "text/csv"),
            "sheet_file": (
                "sheet.csv",
                b"data,descricao,historico,valor\n2026-04-01,RECEBIMENTO,RECEBIMENTO,100",
                "text/csv",
            ),
        },
    )

    assert response.status_code == 422
    assert "ambiguous column mapping" in response.json()["detail"]


def test_reconcile_normalization_preview_aligns_sign_with_same_semantic_description() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": (
                "bank.csv",
                b"date,description,amount\n01/04/2026,pagamento fornecedor alfa,980.00",
                "text/csv",
            ),
            "sheet_file": (
                "sheet.csv",
                b"data,valor,descricao\n2026-04-01,-980.00,PAGAMENTO FORNECEDOR ALFA",
                "text/csv",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    preview = payload["normalization_preview"]
    assert preview[0]["source"] == "bank"
    assert preview[1]["source"] == "sheet"
    assert preview[0]["amount"] == -980.0
    assert preview[1]["amount"] == -980.0
    assert preview[0]["type"] == "outflow"
    assert preview[1]["type"] == "outflow"
    assert payload["exact_matches_count"] == 1
    assert payload["date_tolerance_matches_count"] == 0
    assert payload["description_similarity_matches_count"] == 0
    assert payload["total_matches_count"] == 1
    assert payload["conciliated_count"] == 2
    assert payload["pending_count"] == 0
    assert payload["divergent_count"] == 0
    assert payload["bank_unmatched_count"] == 0
    assert payload["sheet_unmatched_count"] == 0
    assert payload["exact_matches_preview"][0]["match_rule"] == "exact"
    assert payload["problems"] == []
    assert payload["summary"]["conciliated_count"] == 2
    assert payload["summary"]["pending_count"] == 0
    assert payload["summary"]["divergent_count"] == 0


def test_reconcile_matches_with_date_tolerance_plus_or_minus_two_days() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": (
                "bank.csv",
                b"date,description,amount\n2026-04-03,PAGAMENTO FORNECEDOR ALFA,-980.00",
                "text/csv",
            ),
            "sheet_file": (
                "sheet.csv",
                b"data,valor,descricao\n2026-04-01,-980.00,PAGAMENTO FORNECEDOR ALFA",
                "text/csv",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["exact_matches_count"] == 0
    assert payload["date_tolerance_matches_count"] == 1
    assert payload["description_similarity_matches_count"] == 0
    assert payload["total_matches_count"] == 1
    assert payload["conciliated_count"] == 2
    assert payload["pending_count"] == 0
    assert payload["divergent_count"] == 0
    assert payload["bank_unmatched_count"] == 0
    assert payload["sheet_unmatched_count"] == 0
    assert payload["exact_matches_preview"] == []
    assert len(payload["date_tolerance_matches_preview"]) == 1
    assert payload["date_tolerance_matches_preview"][0]["match_rule"] == "date_tolerance"
    assert payload["date_tolerance_matches_preview"][0]["reason"] == "matched_equal_amount_within_2_days"


def test_reconcile_matches_with_description_similarity_when_amount_matches_but_date_is_out_of_tolerance() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": (
                "bank.csv",
                b"date,description,amount\n2026-04-12,PAGAMENTO FORNECEDOR ALFA LTDA,-980.00",
                "text/csv",
            ),
            "sheet_file": (
                "sheet.csv",
                b"data,valor,descricao\n2026-04-01,-980.00,FORNECEDOR ALFA",
                "text/csv",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["exact_matches_count"] == 0
    assert payload["date_tolerance_matches_count"] == 0
    assert payload["description_similarity_matches_count"] == 1
    assert payload["total_matches_count"] == 1
    assert payload["conciliated_count"] == 0
    assert payload["pending_count"] == 0
    assert payload["divergent_count"] == 2
    assert payload["bank_unmatched_count"] == 0
    assert payload["sheet_unmatched_count"] == 0
    assert payload["exact_matches_preview"] == []
    assert payload["date_tolerance_matches_preview"] == []
    assert len(payload["description_similarity_matches_preview"]) == 1
    assert payload["description_similarity_matches_preview"][0]["match_rule"] == "description_similarity"
    assert (
        payload["description_similarity_matches_preview"][0]["reason"]
        == "matched_equal_amount_with_similar_description"
    )
    assert payload["reconciliation_rows"][0]["status"] == "divergente"
    assert payload["reconciliation_rows"][0]["reason"] == "date_out_of_tolerance_window"
    assert payload["reconciliation_rows"][1]["status"] == "divergente"
    assert payload["reconciliation_rows"][1]["reason"] == "date_out_of_tolerance_window"


def test_reconcile_marks_amount_mismatch_as_divergent() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": (
                "bank.csv",
                b"date,description,amount\n2026-04-01,PAGAMENTO FORNECEDOR ALFA,-100.00",
                "text/csv",
            ),
            "sheet_file": (
                "sheet.csv",
                b"data,valor,descricao\n2026-04-02,-120.00,PAGAMENTO FORNECEDOR ALFA",
                "text/csv",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["conciliated_count"] == 0
    assert payload["pending_count"] == 0
    assert payload["divergent_count"] == 2
    assert len(payload["reconciliation_rows"]) == 2
    assert payload["reconciliation_rows"][0]["status"] == "divergente"
    assert payload["reconciliation_rows"][0]["reason"] == "amount_mismatch"
    assert payload["reconciliation_rows"][1]["status"] == "divergente"
    assert payload["reconciliation_rows"][1]["reason"] == "amount_mismatch"
    assert payload["problems"][0]["type"] == "amount_mismatch"


def test_reconcile_marks_date_out_of_tolerance_as_divergent() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": (
                "bank.csv",
                b"date,description,amount\n2026-04-01,PAGAMENTO FORNECEDOR ALFA,-100.00",
                "text/csv",
            ),
            "sheet_file": (
                "sheet.csv",
                b"data,valor,descricao\n2026-04-10,-100.00,PAGAMENTO FORNECEDOR ALFA",
                "text/csv",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["conciliated_count"] == 0
    assert payload["pending_count"] == 0
    assert payload["divergent_count"] == 2
    assert len(payload["reconciliation_rows"]) == 2
    assert payload["reconciliation_rows"][0]["status"] == "divergente"
    assert payload["reconciliation_rows"][0]["reason"] == "date_out_of_tolerance_window"
    assert payload["reconciliation_rows"][1]["status"] == "divergente"
    assert payload["reconciliation_rows"][1]["reason"] == "date_out_of_tolerance_window"


def test_reconcile_generates_operational_problem_insights() -> None:
    client = TestClient(app)

    response = client.post(
        "/reconcile",
        files={
            "bank_file": (
                "bank.csv",
                (
                    b"date,description,amount\n"
                    b"2026-04-01,RECEBIMENTO CLIENTE BETA,300.00\n"
                    b"2026-04-02,PAGAMENTO FORNECEDOR ALFA,-100.00"
                ),
                "text/csv",
            ),
            "sheet_file": (
                "sheet.csv",
                (
                    b"data,valor,descricao\n"
                    b"2026-04-03,-50.00,PAGAMENTO FRETE\n"
                    b"2026-04-02,-120.00,PAGAMENTO FORNECEDOR ALFA"
                ),
                "text/csv",
            ),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    problem_types = {item["type"] for item in payload["problems"]}
    assert "missing_payment" in problem_types
    assert "missing_receipt" in problem_types
    assert "amount_mismatch" in problem_types
    assert payload["summary"] == {
        "total_bank_rows": 2,
        "total_sheet_rows": 2,
        "conciliated_count": 0,
        "pending_count": 2,
        "divergent_count": 2,
    }


def test_reconcile_report_happy_path_csv_and_not_found() -> None:
    client = TestClient(app)
    intake = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,TEST,-100", "text/csv"),
            "sheet_file": ("sheet.csv", b"data,valor,descricao\n2026-04-01,-120,TEST", "text/csv"),
        },
    )
    assert intake.status_code == 200
    analysis_id = intake.json()["analysis_id"]

    xlsx_report = client.get(f"/reconcile-report/{analysis_id}")
    assert xlsx_report.status_code == 200
    assert (
        xlsx_report.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(xlsx_report.content))
    assert "Problemas" in workbook.sheetnames
    problems_sheet = workbook["Problemas"]
    assert [cell.value for cell in problems_sheet[1]] == [
        "row_id",
        "source",
        "date",
        "description",
        "amount",
        "status",
        "reason",
        "matched_row_id",
    ]
    assert problems_sheet.max_row == 3
    assert problems_sheet["F2"].value == "divergente"
    assert problems_sheet["F3"].value == "divergente"
    assert problems_sheet.max_row >= 2

    csv_report = client.get(f"/reconcile-report/{analysis_id}?format=csv")
    assert csv_report.status_code == 200
    assert csv_report.headers["content-type"].startswith("text/csv")
    assert "row_id,source,date,description,amount,status,match_rule,matched_row_id,reason" in csv_report.text

    missing = client.get("/reconcile-report/rc_missing_id")
    assert missing.status_code == 404
    assert missing.json()["detail"] == "Analysis not found"


def test_reconcile_report_includes_fallback_problem_row_when_no_issues() -> None:
    client = TestClient(app)
    intake = client.post(
        "/reconcile",
        files={
            "bank_file": ("bank.csv", b"date,description,amount\n2026-04-01,TEST,-100", "text/csv"),
            "sheet_file": ("sheet.csv", b"data,valor,descricao\n2026-04-01,-100,TEST", "text/csv"),
        },
    )
    assert intake.status_code == 200
    analysis_id = intake.json()["analysis_id"]

    xlsx_report = client.get(f"/reconcile-report/{analysis_id}")
    assert xlsx_report.status_code == 200
    workbook = load_workbook(filename=BytesIO(xlsx_report.content))
    problems_sheet = workbook["Problemas"]

    assert [cell.value for cell in problems_sheet[1]] == [
        "row_id",
        "source",
        "date",
        "description",
        "amount",
        "status",
        "reason",
        "matched_row_id",
    ]
    assert [cell.value for cell in problems_sheet[2]] == [
        "none",
        "system",
        None,
        "No pending/divergent issues were detected.",
        None,
        "none",
        "none",
        None,
    ]
