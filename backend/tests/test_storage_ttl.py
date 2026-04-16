from datetime import datetime, timedelta, timezone

import pytest
from openpyxl import load_workbook

from app.application import AnalysisNotFoundError, TempAnalysisStorage
from app.application.models import AnalysisData, TransactionRow
from app.application.ofx_parser import parse_ofx_transactions


def _build_analysis_data(analysis_id: str = "an_testttl") -> AnalysisData:
    return AnalysisData(
        analysis_id=analysis_id,
        file_type="csv",
        upload_filename="sample.csv",
        transactions_total=1,
        total_inflows=100.0,
        total_outflows=-20.0,
        net_total=80.0,
        preview_transactions=[
            TransactionRow(
                date="2026-04-01",
                description="TEST",
                amount=-20.0,
                category="Outros",
                reconciliation_status="unmatched",
            )
        ],
    )


def test_save_analysis_persists_created_and_expires_at(tmp_path) -> None:
    base_now = datetime(2026, 4, 10, 12, 0, tzinfo=timezone.utc)
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600, now_provider=lambda: base_now)

    storage.save_analysis(_build_analysis_data())

    analysis_json = tmp_path / "an_testttl" / "analysis.json"
    content = analysis_json.read_text(encoding="utf-8")

    assert '"created_at": "2026-04-10T12:00:00+00:00"' in content
    assert '"expires_at": "2026-04-10T13:00:00+00:00"' in content


def test_get_report_path_returns_report_when_analysis_is_not_expired(tmp_path) -> None:
    current = {"now": datetime(2026, 4, 10, 12, 0, tzinfo=timezone.utc)}
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=7200, now_provider=lambda: current["now"])

    storage.save_analysis(_build_analysis_data())
    current["now"] = current["now"] + timedelta(minutes=30)

    report_path = storage.get_report_path("an_testttl")
    assert report_path.exists()


def test_get_report_path_raises_not_found_and_cleans_directory_when_expired(tmp_path) -> None:
    current = {"now": datetime(2026, 4, 10, 12, 0, tzinfo=timezone.utc)}
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=60, now_provider=lambda: current["now"])

    storage.save_analysis(_build_analysis_data())
    current["now"] = current["now"] + timedelta(minutes=2)

    with pytest.raises(AnalysisNotFoundError):
        storage.get_report_path("an_testttl")

    assert not (tmp_path / "an_testttl").exists()


def test_save_analysis_formats_report_for_readability(tmp_path) -> None:
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600)
    storage.save_analysis(_build_analysis_data())

    report_path = tmp_path / "an_testttl" / "report.xlsx"
    workbook = load_workbook(report_path)
    sheet = workbook["Transacoes"]

    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:E2"
    assert sheet.column_dimensions["B"].width is not None
    assert sheet.column_dimensions["B"].width >= 12
    assert "Conciliacao" in workbook.sheetnames


def test_save_analysis_writes_full_transaction_set_to_report(tmp_path) -> None:
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600)
    full_rows = [
        TransactionRow(
            date=f"2026-04-{index:02d}",
            description=f"TX {index}",
            amount=float(-index),
            category="Outros",
            reconciliation_status="unmatched",
        )
        for index in range(1, 26)
    ]
    data = AnalysisData(
        analysis_id="an_full_report",
        file_type="csv",
        upload_filename="full.csv",
        transactions_total=len(full_rows),
        total_inflows=0.0,
        total_outflows=-325.0,
        net_total=-325.0,
        preview_transactions=full_rows[:20],
        report_transactions=full_rows,
    )

    storage.save_analysis(data)

    report_path = tmp_path / "an_full_report" / "report.xlsx"
    workbook = load_workbook(report_path)
    sheet = workbook["Transacoes"]
    assert sheet.max_row == 26


def test_save_analysis_writes_convert_ofx_artifact(tmp_path) -> None:
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600)
    storage.save_analysis(_build_analysis_data())

    ofx_path = tmp_path / "an_testttl" / "converted.ofx"
    assert ofx_path.exists()

    parsed = parse_ofx_transactions(ofx_path.read_bytes())
    assert len(parsed) == 1
    assert parsed[0].description == "TEST"


def test_save_analysis_writes_convert_csv_artifact(tmp_path) -> None:
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600)
    storage.save_analysis(_build_analysis_data())

    csv_path = tmp_path / "an_testttl" / "converted.csv"
    assert csv_path.exists()
    assert "date,description,amount,category,reconciliation_status" in csv_path.read_text(encoding="utf-8")
